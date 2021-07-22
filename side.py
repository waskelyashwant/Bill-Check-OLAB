from flask import Flask , render_template , request
from flask import send_file
from selenium import webdriver
import selenium
from selenium.webdriver.remote import webelement
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.keys import Keys
import openpyxl
import os
import time
import pandas as pd
import threading
import sys


chrome_options = webdriver.ChromeOptions()
chrome_options.binary_location = os.environ.get("GOOGLE_CHROME_BIN")
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.add_argument("--no-sandbox")


def amazon_zone(k_no, driver,ags):
    try:
        link = "https://www.amazon.in/hfc/bill/electricity?ref_=apay_deskhome_Electricity"
        driver.get(link)
        driver.find_element_by_class_name("a-dropdown-prompt").click()
        x = driver.find_element_by_class_name("a-dropdown-common")
        ul = x.find_element_by_tag_name("ul")
        li = ul.find_elements_by_tag_name("li")
        li[26].click()
        time.sleep(2)
        x1 =Select(driver.find_element_by_id("ELECTRICITY>hfc-states-rajasthan"))
        # x1.select_by_index(1)
        # x1._setSelected
        # webdriver.ActionChains(driver).send_keys(Keys.ENTER).perform()
        # time.sleep(2)
        # x1 =driver.find_element_by_id("ELECTRICITY>hfc-states-rajasthan").click()
        x = driver.find_elements_by_class_name("a-dropdown-prompt")
        x[1].click()
        ul=driver.find_element_by_class_name("a-box-list")
        li=ul.find_elements_by_tag_name("li")
        li[0].click() 
        time.sleep(4)
        # ags = sheet.cell(row=index, column=5).value
        pk = "Continue to Pay ₹"+ str(ags)+".00"
        driver.find_element_by_id("K Number").click()
        driver.find_element_by_id("K Number").clear()
        driver.find_element_by_id("K Number").send_keys(k_no)
        driver.find_element_by_id("fetchBtnText").click()
        daa = driver.find_element_by_id("paymentBtnAmountText")
        print(daa.get_attribute("innerHTML"))
        if pk==daa.get_attribute("innerHTML"):
            return "unpaid"
            # sheet.cell(row = index, column = 6).value = "unpaid"
            # df.save('status.xlsx')
        else:
            return "paid"
            # print("asd")
            # sheet.cell(row = index, column = 6).value = "paid"
    except:
        return "Unable to check"
    
    time.sleep(2)


def jodhpur_zone(k_no, driver):
    try:
        print("Inside Jodhpur")
        link = "http://wss.rajdiscoms.com/HDFC_QUICKPAY/index"
        driver.get(link)
        driver.find_element_by_id("txtKno").click()
        driver.find_element_by_id("txtKno").clear()
        driver.find_element_by_id("txtKno").send_keys(k_no)
        driver.find_element_by_id("txtEmail").click()
        driver.find_element_by_id("txtEmail").clear()
        driver.find_element_by_id("txtEmail").send_keys("admin@gmail.com")
        driver.find_element_by_id("btnsearch").click()
        status = driver.find_element_by_id("lblMessage").text
        print(k_no, status)
        return status
    except:
        print("unable to check")
        return "Unable to check"
    # sheet.cell(row = index, column = 6).value = status


def ajmer_zone(index, k_no, driver, sheet):
    link = "https://jansoochna.rajasthan.gov.in/Services/DynamicControls"
    driver.get(link)
    driver.find_element_by_partial_link_text("Know about your Electricity Bill Payment Information - AVVNL").click()
    time.sleep(2)
    driver.find_element_by_id("Enter_your_K_number").click()
    driver.find_element_by_id("Enter_your_K_number").clear()
    driver.find_element_by_id("Enter_your_K_number").send_keys(k_no)  
    driver.find_element_by_id("btnSubmit").click()
    time.sleep(2)   
    amnt= driver.find_element_by_xpath("/html/body/div[1]/section/div[3]/div/div/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr[1]/td[9]")
    
    # print(asds.get_attribute("innerHTML"))
    dat =  driver.find_element_by_xpath("/html/body/div[1]/section/div[3]/div/div/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr[1]/td[12]")
    if amnt==sheet.cell(row=index, column=4).value:
        sheet.cell(row = index, column = 6).value = "paid"
        df.save('status.xlsx')
    else:
        sheet.cell(row = index, column = 6).value = "unpaid"
        df.save('status.xlsx')


def jansoochna_zone(index, k_no, driver, sheet):
    link = "https://jansoochna.rajasthan.gov.in/Services/DynamicControls"
    driver.get(link)
    driver.find_element_by_partial_link_text("Know about your Electricity Bill Payment Information - JDVVNL").click()
    time.sleep(2)
    driver.find_element_by_id("Enter_your_K_number").click()
    driver.find_element_by_id("Enter_your_K_number").clear()
    driver.find_element_by_id("Enter_your_K_number").send_keys(k_no.value)  
    driver.find_element_by_id("btnSubmit").click()
    time.sleep(2)   
    amnt= driver.find_element_by_xpath("/html/body/div[1]/section/div[3]/div/div/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr[1]/td[12]")
    amnt = amnt.get_attribute("innerHTML")
    print(amnt)
    dat =  driver.find_element_by_xpath("/html/body/div[1]/section/div[3]/div/div/div/div/div[3]/div[2]/div/div[1]/div[2]/table/tbody/tr[1]/td[13]")
    dat = dat.get_attribute("innerHTML")
    if amnt==str(sheet.cell(row=index, column=4).value):
        sheet.cell(row = index, column = 6).value = "paid"
    else:
        sheet.cell(row = index, column = 6).value = "unpaid"



def starting(real_list, lista, result, mapping_dict):
    driver = webdriver.Chrome(executable_path=os.environ.get("CHROMEDRIVER_PATH"), options=chrome_options)
    # driver=webdriver.Chrome("chromedriver.exe")
    x=len(lista)
    for k in range(lista[0],lista[1]):
        distr = real_list[k][1].value
        zone=mapping_dict[distr]
        k_no=real_list[k][3].value
        print(distr, zone, k_no)
        # time.sleep(5)
        if zone=='Jodhpur':
            # print("Jodhpur")
            # driver = webdriver.Chrome(executable_path=os.environ.get("CHROMEDRIVER_PATH"), chrome_options=chrome_options)
            # driver=webdriver.Chrome("chromedriver.exe")
            status = jodhpur_zone(k_no,driver)
            # driver.close()
            print("Processed")

        elif zone=='Jaipur' or zone=="Bharatpur" or zone=="Bikaner" or zone=="Kota" or zone=="TP":
            ags = real_list[k][4].value
            status = amazon_zone(k_no,driver,ags)
        elif zone=='Ajmer':
            status = ajmer_zone(k_no, driver)
        else:
            status = jansoochna_zone(k_no, driver)
        result.append(status)
    driver.close()


data = pd.read_excel(r'mapping.xlsx')
df1 = pd.DataFrame(data)
length = len(df1.index)
mapping_dict={}
for i in range(0,length):
    mapping_dict[df1.iloc[i]['Distributor as per file']]=str(df1.iloc[i]['Zone'])
print(mapping_dict)

# print(f)  
f="bill_file.xlsx"

df = openpyxl.load_workbook(f)
sheet = df.active

z=0
real_list=[]
for k in sheet:
    if z==0:
        z+=1
        continue
    k=list(k)
    if k[1].value==None:
        continue
    real_list.append(k)
x=len(real_list)
# y=int(x/20)
# main_list=[[0, y],[y, 2*y],[2*y, 3*y],[3*y, 4*y],[4*y, 5*y],[5*y, 6*y],[6*y, 7*y],[7*y, 8*y],[8*y, 9*y],[9*y, 10*y],
#            [10*y, 11*y],[11*y, 12*y],[12*y, 13*y],[13*y, 14*y],[14*y, 15*y],[15*y, 16*y],[16*y, 17*y],[17*y, 18*y],[18*y, 19*y],[19*y, 20*y+ x %20]]

y=int(x/3)
main_list=[[0,y], [y, 2*y], [2*y, 3*y+x%3]]

threads=[]
results=[]
for i in range(0,3):
    res=[None]*(y+x%3)
    results.append(res)

for v in range(0,3):
    t=threading.Thread(target=starting, args=(real_list,main_list[v],results[v], mapping_dict))
    t.start()
    threads.append(t)
for v in threads:
    v.join()

print(results)
main_result=[]
for v in range(0,3):
    length = main_list[v][1] - main_list[v][0]
    flag=0
    for u in range(0,len(results[v])):
        if results[v][u]!=None:
            flag=1
            main_result.append(results[v][u])
    if flag==0:
        for i in range(0,length):
            main_result.append("crashed")

index=1
sheet.cell(row = index, column = 6).value = 'Status'
df.save('status.xlsx')
index+=1
for k in range(0,len(main_result)):
    if sheet.cell(row=index, column= 1).value == None:
        break
    sheet.cell(row = index, column = 6).value = main_result[k]
    index+=1
    df.save('status.xlsx')

file1 = open("value.txt","w")
file1.write("1")
file1.close() 
